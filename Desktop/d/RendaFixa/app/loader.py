from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional
from openpyxl import load_workbook

from app.models import Titulo

_FILE_PATH = Path(__file__).parent.parent / "taxas_indicativas.xlsm"


def load_titulos(path: Path = _FILE_PATH) -> Dict[int, Titulo]:
    if not path.exists():
        raise FileNotFoundError(f"Arquivo de títulos não encontrado: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    rows = [*parse_bancario(wb), *parse_publico(wb)]

    for idx, titulo in enumerate(rows, start=1):
        titulo.id = idx

    return {t.id: t for t in rows}


def excel_date(value) -> Optional[str]:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(value))).strftime("%Y-%m-%d")
        except Exception:
            return None
    return None


def to_float(value) -> Optional[float]:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_bancario(wb) -> List[Titulo]:
    ws = wb["Crédito bancário"]
    titulos = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if row[0] == "Emissor":
            header_found = True
            continue
        if not header_found or not row[0]:
            continue

        titulos.append(Titulo(
            id=0,
            emissor=str(row[0]),
            produto=str(row[1]) if row[1] else "",
            prazo_dias=int(row[2]) if isinstance(row[2], (int, float)) else None,
            vencimento=excel_date(row[3]),
            indexador=str(row[4]) if row[4] else "",
            taxa_maxima=to_float(row[5]),
            taxa_portal=to_float(row[6]),
            spread_portal=to_float(row[7]),
            receita_estimada=to_float(row[8]),
            aplicacao_minima=to_float(row[9]),
            rating=str(row[10]) if row[10] else None,
            juros=str(row[11]) if row[11] else None,
            fonte="bancario",
        ))
    return titulos


def parse_publico(wb) -> List[Titulo]:
    ws = wb["Títulos Públicos"]
    titulos = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if row[0] == "Título":
            header_found = True
            continue
        if not header_found or not row[0]:
            continue

        titulos.append(Titulo(
            id=0,
            emissor="Tesouro Nacional",
            produto=str(row[0]),
            prazo_dias=None,
            vencimento=excel_date(row[1]),
            indexador=str(row[0]),
            taxa_maxima=to_float(row[2]),
            taxa_portal=to_float(row[2]),
            spread_portal=to_float(row[3]),
            receita_estimada=to_float(row[4]),
            aplicacao_minima=30.0,
            rating="AAA",
            juros=None,
            fonte="publico",
        ))
    return titulos
